# Copyright (c) 2026
# Attendance Report - server side logic

BRANCH_WARD_FIELD = "custom_ward"

import json
from io import BytesIO

import frappe
from frappe import _
from frappe.utils import add_days, formatdate, getdate

PREFIX_COLUMNS = (
    {"label": _("Ward"), "fieldname": "ward", "fieldtype": "Data", "width": 100},
    {"label": _("Organization"), "fieldname": "branch", "fieldtype": "Link", "options": "Branch", "width": 140},
    {"label": _("Employee ID"), "fieldname": "employee", "fieldtype": "Link", "options": "Employee", "width": 110},
    {"label": _("Employee Name"), "fieldname": "employee_name", "fieldtype": "Data", "width": 160},
    {"label": _("Department"), "fieldname": "department", "fieldtype": "Link", "options": "Department", "width": 130},
    {"label": _("Designation"), "fieldname": "designation", "fieldtype": "Data", "width": 110},
)

SUFFIX_COLUMNS = (
    {"label": _("Total Present"), "fieldname": "total_present", "fieldtype": "Int", "width": 90},
    {"label": _("Total Absent"), "fieldname": "total_absent", "fieldtype": "Int", "width": 90},
    {"label": _("Total Leave"), "fieldname": "total_leave", "fieldtype": "Int", "width": 90},
    {"label": _("Remarks"), "fieldname": "remarks", "fieldtype": "Data", "width": 120},
)


def execute(filters=None):
    filters = frappe._dict(filters or {})
    filters.branch = parse_list(filters.get("branch"))
    filters.ward = parse_list(filters.get("ward"))
    filters.department = parse_list(filters.get("department"))

    validate_filters(filters)

    dates = get_dates_in_range(filters.from_date, filters.to_date)
    columns = get_columns(dates)
    data = get_data(filters, dates)
    return columns, data


def parse_list(value):
    if not value:
        return []
    if isinstance(value, str):
        value = frappe.parse_json(value)
    return value or []


def _strip_quotes(value):
    if value is None:
        return value
    if isinstance(value, str):
        v = value.strip()
        if len(v) >= 2 and v.startswith('"') and v.endswith('"'):
            return v[1:-1]
    return value


def validate_filters(filters):
    if not filters.get("from_date") or not filters.get("to_date"):
        frappe.throw(_("From Date and To Date are mandatory"))

    if getdate(filters.from_date) > getdate(filters.to_date):
        frappe.throw(_("From Date cannot be after To Date"))

    if filters.department and not filters.branch:
        frappe.throw(_("Please select at least one Organization (Branch) before filtering by Department"))

    if not filters.branch and not filters.ward:
        frappe.throw(_("Please select a Ward or Organization (Branch) from the filter to generate this report"))

    from frappe.utils import date_diff
    if date_diff(filters.to_date, filters.from_date) > 90:
        frappe.throw(_("Date range cannot exceed 90 days for this report"))


def get_dates_in_range(from_date, to_date):
    dates = []
    current = getdate(from_date)
    end = getdate(to_date)
    while current <= end:
        dates.append(current)
        current = add_days(current, 1)
    return dates


def get_columns(dates):
    columns = list(PREFIX_COLUMNS)
    for d in dates:
        columns.append(
            {"label": str(d.day), "fieldname": f"d_{d.isoformat()}", "fieldtype": "Data", "width": 42}
        )
    columns.extend(SUFFIX_COLUMNS)
    return columns


def leave_abbreviation(leave_type):
    if not leave_type:
        return "L"
    return "".join(word[0] for word in leave_type.split() if word).upper()


def day_value(status, leave_type):
    if status in ("Present", "Work From Home"):
        return "P"
    if status == "Half Day":
        return "HD"
    if status == "Absent":
        return "A"
    if status == "On Leave":
        return leave_abbreviation(leave_type)
    return ""


# ---------------------------------------------------------------------------
# Strict Permitted Employee Extraction
# ---------------------------------------------------------------------------

def _get_permitted_employees():
    """
    Returns ONLY the exact employee records visible in the logged-in user's
    standard Employee List View.
    """
    if frappe.session.user == "Administrator":
        return None

    # Evaluates native Role & User Permissions strictly without manual expansion
    return frappe.get_all("Employee", pluck="name", ignore_permissions=False)


def get_attendance_records(filters):
    Attendance = frappe.qb.DocType("Attendance")
    Employee = frappe.qb.DocType("Employee")
    Branch = frappe.qb.DocType("Branch")

    # 1. Strictly restrict to permitted employees first
    permitted_employees = _get_permitted_employees()
    if permitted_employees is not None and not permitted_employees:
        return []

    query = (
        frappe.qb.from_(Attendance)
        .inner_join(Employee)
        .on(Attendance.employee == Employee.name)
        .left_join(Branch)
        .on(Employee.branch == Branch.name)
        .select(
            Attendance.employee,
            Attendance.employee_name,
            Attendance.attendance_date,
            Attendance.status,
            Attendance.leave_type,
            Employee.branch,
            Employee.department,
            Employee.designation,
            Branch[BRANCH_WARD_FIELD].as_("ward"),
        )
        .where(Attendance.docstatus == 1)
        .where(Attendance.attendance_date >= filters.from_date)
        .where(Attendance.attendance_date <= filters.to_date)
    )

    if permitted_employees is not None:
        query = query.where(Employee.name.isin(permitted_employees))

    if filters.branch:
        query = query.where(Employee.branch.isin(filters.branch))
    if filters.ward:
        query = query.where(Branch[BRANCH_WARD_FIELD].isin(filters.ward))
    if filters.department:
        query = query.where(Employee.department.isin(filters.department))

    query = query.orderby(Attendance.employee).orderby(Attendance.attendance_date)

    return query.run(as_dict=True)


def get_data(filters, dates):
    records = get_attendance_records(filters)

    grouped = {}
    for rec in records:
        date = getdate(rec.attendance_date)
        group = grouped.setdefault(
            rec.employee,
            {
                "ward": rec.ward,
                "branch": rec.branch,
                "employee": rec.employee,
                "employee_name": rec.employee_name,
                "department": rec.department,
                "designation": rec.designation,
                "days": {},
                "leave_counts": {},
            },
        )
        value = day_value(rec.status, rec.leave_type)
        group["days"][date.isoformat()] = value

        if rec.status == "On Leave":
            abbr = leave_abbreviation(rec.leave_type)
            group["leave_counts"][abbr] = group["leave_counts"].get(abbr, 0) + 1

    groups = sorted(grouped.values(), key=lambda g: (g["employee_name"] or "", g["employee"]))

    data = []
    for group in groups:
        row = {
            "ward": _strip_quotes(group["ward"]),
            "branch": _strip_quotes(group["branch"]),
            "employee": _strip_quotes(group["employee"]),
            "employee_name": group["employee_name"],
            "department": _strip_quotes(group["department"]),
            "designation": group["designation"],
        }

        total_present = total_absent = total_leave = 0
        for d in dates:
            value = group["days"].get(d.isoformat(), "")
            row[f"d_{d.isoformat()}"] = value
            if value in ("P", "HD"):
                total_present += 1
            elif value == "A":
                total_absent += 1
            elif value:
                total_leave += 1

        row["total_present"] = total_present
        row["total_absent"] = total_absent
        row["total_leave"] = total_leave

        if group["leave_counts"]:
            row["remarks"] = ", ".join(
                f"{count} {abbr}" for abbr, count in sorted(group["leave_counts"].items())
            )
        else:
            row["remarks"] = "-"

        data.append(row)

    return data


@frappe.whitelist()
def download_excel(filters=None):
    if isinstance(filters, str):
        filters = json.loads(filters)
    filters = frappe._dict(filters or {})
    filters.branch = parse_list(filters.get("branch"))
    filters.ward = parse_list(filters.get("ward"))
    filters.department = parse_list(filters.get("department"))
    validate_filters(filters)

    dates = get_dates_in_range(filters.from_date, filters.to_date)
    columns = get_columns(dates)
    data = get_data(filters, dates)

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    groups = []
    current = None
    for d in dates:
        key = (d.year, d.month)
        if current is None or current["key"] != key:
            current = {"key": key, "year": d.year, "month": d.month, "count": 1, "start_col": None, "end_col": None}
            groups.append(current)
        else:
            current["count"] += 1

    prefix_count = len(PREFIX_COLUMNS)
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center")
    month_palette = [
        "1f77b4", "ff7f0e", "2ca02c", "d62728",
        "9467bd", "8c564b", "e377c2", "7f7f7f",
        "bcbd22", "17becf", "393b79", "637939",
    ]
    month_names = [
        "January", "February", "March", "April", "May", "June",
        "July", "August", "September", "October", "November", "December",
    ]

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=prefix_count)
    identity_cell = ws.cell(1, 1, "")
    identity_cell.fill = header_fill
    identity_cell.font = header_font
    identity_cell.alignment = center
    identity_cell.border = border

    for i, g in enumerate(groups):
        start_col = prefix_count + 1 + sum(x["count"] for x in groups[:i])
        end_col = start_col + g["count"] - 1
        colour = month_palette[(g["month"] - 1) % len(month_palette)]
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        cell = ws.cell(1, start_col, f"{month_names[g['month'] - 1]} {g['year']}")
        cell.fill = PatternFill(start_color=colour, end_color=colour, fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = center
        cell.border = border

    trailing_count = len(SUFFIX_COLUMNS)
    trailing_start_col = prefix_count + len(dates) + 1
    ws.merge_cells(
        start_row=1, start_column=trailing_start_col,
        end_row=1, end_column=trailing_start_col + trailing_count - 1,
    )
    trailing_header = ws.cell(1, trailing_start_col, "")
    trailing_header.fill = header_fill
    trailing_header.font = header_font
    trailing_header.alignment = center
    trailing_header.border = border

    for col_idx, col in enumerate(columns, start=1):
        cell = ws.cell(2, col_idx, col["label"])
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 26
    ws.freeze_panes = "A3"

    for row_idx, row in enumerate(data, start=3):
        for col_idx, col in enumerate(columns, start=1):
            value = row.get(col["fieldname"], "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            if col["fieldname"] in {"employee_name", "ward", "remarks"}:
                cell.alignment = left_align
            else:
                cell.alignment = center

    for col_idx, col in enumerate(columns, start=1):
        width = col.get("width") or 100
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(int(width / 7), 8), 32)

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_options.horizontalCentered = True

    buffer = BytesIO()
    wb.save(buffer)

    frappe.local.response.filename = (
        f"Attendance Report - {formatdate(filters.from_date, 'dd-MM-yyyy')}"
        f" to {formatdate(filters.to_date, 'dd-MM-yyyy')}.xlsx"
    )
    frappe.local.response.filecontent = buffer.getvalue()
    frappe.local.response.type = "binary"


# ---------------------------------------------------------------------------
# Filter Whitelists (Scoped strictly to visible employees)
# ---------------------------------------------------------------------------

def _get_permitted_branches():
    permitted_employees = _get_permitted_employees()
    if permitted_employees is None:
        return [b.name for b in frappe.get_all("Branch", order_by="name asc")]
    if not permitted_employees:
        return []

    branches = frappe.get_all(
        "Employee",
        filters={"name": ("in", permitted_employees)},
        pluck="branch",
        distinct=True,
    )
    return sorted({b for b in branches if b})


@frappe.whitelist()
def get_permitted_branch_options(txt="", branches=None):
    permitted = _get_permitted_branches()
    result = permitted

    selected = parse_list(branches)
    if selected:
        result = [b for b in result if b in selected]

    if txt:
        needle = txt.strip().lower()
        result = [b for b in result if needle in b.lower()]

    return result


@frappe.whitelist()
def get_permitted_ward_options(txt="", branches=None):
    permitted_branches = _get_permitted_branches()
    if not permitted_branches:
        return []

    selected = parse_list(branches)
    scope_branches = [b for b in permitted_branches if not selected or b in selected]

    if not scope_branches:
        return []

    filters = {"name": ("in", scope_branches)}
    if txt:
        filters[BRANCH_WARD_FIELD] = ["like", f"%{txt}%"]

    wards = frappe.get_all(
        "Branch",
        filters=filters,
        pluck=BRANCH_WARD_FIELD,
        distinct=True,
    )
    return sorted({w for w in wards if w})


@frappe.whitelist()
def get_permitted_department_options(txt="", branches=None):
    permitted_employees = _get_permitted_employees()
    if permitted_employees is not None and not permitted_employees:
        return []

    emp_filters = {}
    if permitted_employees is not None:
        emp_filters["name"] = ("in", permitted_employees)

    selected = parse_list(branches)
    if selected:
        emp_filters["branch"] = ("in", selected)

    departments = frappe.get_all(
        "Employee",
        filters=emp_filters,
        pluck="department",
        distinct=True,
    )

    result = sorted({d for d in departments if d})

    if txt:
        needle = txt.strip().lower()
        result = [d for d in result if needle in d.lower()]

    return result