# Copyright (c) 2026
# Attendance Source Report.py - server side logic
#
# Fieldname configuration:
BRANCH_WARD_FIELD = "custom_ward"      # Ward field on Branch doctype
EMPLOYEE_HOD_FIELD = "custom_hod"             # HOD field on Employee (stores Employee ID of HOD)
CHECKIN_SOURCE_FIELD = "custom_source" # Source field on Employee Checkin (Data)
# "Biometric" -> Biometric, empty/null -> Web,
# value containing "mobile" -> Mobile App

import json
from io import BytesIO

import frappe
from frappe import _
from frappe.utils import getdate, formatdate


def execute(filters=None):
	filters = frappe._dict(filters or {})
	validate_filters(filters)

	columns = get_columns()
	data = get_data(filters)
	return columns, data


def validate_filters(filters):
	if not filters.from_date or not filters.to_date:
		frappe.throw(_("From Date and To Date are mandatory"))

	if getdate(filters.from_date) > getdate(filters.to_date):
		frappe.throw(_("From Date cannot be after To Date"))

	# Department filter is only valid when Organization (Branch) is chosen
	if filters.department and not filters.branch:
		frappe.throw(_("Please select an Organization (Branch) before filtering by Department"))

	# Require at least Ward or Branch to scope the report
	ward_values = _as_list(filters.get("ward"))
	branch_values = _as_list(filters.get("branch"))
	if not ward_values and not branch_values:
		frappe.throw(_("Please select a Ward or Organization (Branch) from the filter to generate this report"))


def get_columns():
	return [
		{"label": _("Ward Name"), "fieldname": "ward", "fieldtype": "Data", "width": 120},
		{"label": _("Organization (Branch)"), "fieldname": "branch", "fieldtype": "Link", "options": "Branch", "width": 150},
		{"label": _("Employee ID"), "fieldname": "employee", "fieldtype": "Data", "width": 110},
		{"label": _("Employee Name"), "fieldname": "employee_name", "fieldtype": "Data", "width": 180},
		{"label": _("Department Name"), "fieldname": "department", "fieldtype": "Link", "options": "Department", "width": 140},
		{"label": _("Attendance Date"), "fieldname": "attendance_date", "fieldtype": "Date", "width": 110},
		{"label": _("HOD"), "fieldname": "hod_name", "fieldtype": "Data", "width": 150},
		{"label": _("Check-In Source"), "fieldname": "check_in_source", "fieldtype": "Data", "width": 110},
		{"label": _("Check-Out Source"), "fieldname": "check_out_source", "fieldtype": "Data", "width": 110},
		{"label": _("Attendance"), "fieldname": "attendance", "fieldtype": "Data", "width": 100},
		{"label": _("Shift Time"), "fieldname": "shift_time", "fieldtype": "Data", "width": 120},
		{"label": _("Check-In Time"), "fieldname": "check_in_time", "fieldtype": "Data", "width": 100},
		{"label": _("Check-Out Time"), "fieldname": "check_out_time", "fieldtype": "Data", "width": 100},
		{"label": _("Working Hours"), "fieldname": "working_hours", "fieldtype": "Data", "width": 110},
		{"label": _("Late Punch"), "fieldname": "late_punch", "fieldtype": "Data", "width": 90},
		{"label": _("Early Out"), "fieldname": "early_out", "fieldtype": "Data", "width": 90},
		{"label": _("Missed Punch"), "fieldname": "missed_punch", "fieldtype": "Data", "width": 100},
	]


def get_data(filters):
	# CHANGED: rows are now driven by the Attendance list.
	# One report row per submitted Attendance record in the date range.
	attendance_records = get_attendance_records(filters)
	if not attendance_records:
		return []

	employee_ids = list({rec.employee for rec in attendance_records})

	checkin_map = get_checkin_map(filters, employee_ids)
	shift_map = get_shift_map()

	data = []
	for att in attendance_records:
		# Strip stray ASCII double-quotes that some upstream imports
		# embedded into the employee value (e.g. "HR-EMP-03603" -> HR-EMP-03603).
		clean_employee = _clean_employee_value(att.employee)
		key = (clean_employee, getdate(att.attendance_date))
		checkins = checkin_map.get(key, [])

		row = build_row(att, checkins, shift_map)
		row["employee"] = clean_employee
		data.append(row)

	return data


def get_attendance_records(filters):
	# CHANGED: was get_employees(). Now the query starts from Attendance and
	# joins Employee / Branch / HOD to fill the linked-doctype columns.
	# All filters (branch, ward, department) apply to these attendance rows.
	# Branch and Ward are now multi-select lists; we resolve them to actual
	# branch names in one query to keep the attendance join small.
	Attendance = frappe.qb.DocType("Attendance")
	Employee = frappe.qb.DocType("Employee")
	Branch = frappe.qb.DocType("Branch")
	# Self-join on Employee to resolve HOD's employee ID -> HOD's name
	HOD = frappe.qb.DocType("Employee").as_("hod_emp")

	query = (
		frappe.qb.from_(Attendance)
		.inner_join(Employee)
		.on(Attendance.employee == Employee.name)
		.left_join(Branch)
		.on(Employee.branch == Branch.name)
		.left_join(HOD)
		.on(Employee[EMPLOYEE_HOD_FIELD] == HOD.name)
		.select(
			Attendance.employee,
			Attendance.employee_name,
			Attendance.attendance_date,
			Attendance.status,
			Attendance.shift,
			Attendance.in_time,
			Attendance.out_time,
			Attendance.late_entry,
			Attendance.early_exit,
			Attendance.custom_late_arrival_minutes,
			Attendance.custom_early_out_minutes,
			Employee.branch,
			Employee.department,
			Employee.default_shift,
			Branch[BRANCH_WARD_FIELD].as_("ward"),
			HOD.employee_name.as_("hod_name"),
		)
		.where(Attendance.docstatus == 1)
		.where(Attendance.attendance_date >= filters.from_date)
		.where(Attendance.attendance_date <= filters.to_date)
		.orderby(Attendance.attendance_date)
		.orderby(Attendance.employee)
	)

	# Department alone is fine if the user supplied it (validated upstream).
	# Branch+Ward: if only Ward(s) supplied -> resolve which branches to filter on.
	# If Ward is empty but Branch is supplied -> use branches directly.
	ward_values = _as_list(filters.get("ward"))
	branch_values = _as_list(filters.get("branch"))
	department_values = _as_list(filters.get("department"))

	if ward_values and not branch_values:
		resolved_branches = frappe.get_all(
			"Branch",
			filters={BRANCH_WARD_FIELD: ["in", ward_values]},
			pluck="name",
		)
		if resolved_branches:
			query = query.where(Employee.branch.isin(resolved_branches))
		else:
			# No branch matches the chosen ward(s) -> no data possible
			return []
	elif branch_values:
		query = query.where(Employee.branch.isin(branch_values))

	if department_values:
		# Department options are already filtered on the client to branches
		# belonging to selected ward(s)/org(s), so an IN clause is sufficient.
		query = query.where(Employee.department.isin(department_values))

	return query.run(as_dict=True)


def _strip_quotes(value):
	"""Remove stray ASCII double-quotes wrapping a value.

	Some upstream imports stored employee/branch/department IDs wrapped in
	literal '"..."' characters. Strip them so the report renders cleanly.
	"""
	if value is None:
		return value
	if isinstance(value, str):
		v = value.strip()
		if len(v) >= 2 and v.startswith('"') and v.endswith('"'):
			return v[1:-1]
	return value


def _clean_employee_value(value):
	"""Strip quotes from an employee field value if present."""
	return _strip_quotes(value)


def _as_list(value):
	"""Coerce a filter value (str, list, tuple, None) to a flat list of strings."""
	if not value:
		return []
	if isinstance(value, (list, tuple)):
		return [str(v) for v in value if v not in (None, "")]
	if isinstance(value, str):
		# Could be a JSON array from client MultiSelect (with double-quoted strings)
		# or a single value
		v = value.strip()
		if v.startswith("[") and v.endswith("]"):
			try:
				parsed = json.loads(v)
				if isinstance(parsed, list):
					return [str(x) for x in parsed if x not in (None, "")]
			except Exception:
				pass
			# Common client quirk: Python repr ['a','b'] instead of JSON ["a","b"]
			try:
				import ast as _ast
				parsed = _ast.literal_eval(v)
				if isinstance(parsed, (list, tuple)):
					return [str(x) for x in parsed if x not in (None, "")]
			except Exception:
				pass
		return [v]
	return [str(value)]


def get_checkin_map(filters, employee_ids):
	checkins = frappe.get_all(
		"Employee Checkin",
		filters={
			"employee": ["in", employee_ids],
			"time": [
				"between",
				[f"{filters.from_date} 00:00:00", f"{filters.to_date} 23:59:59"],
			],
		},
		fields=["employee", "time", "log_type", CHECKIN_SOURCE_FIELD],
		order_by="time asc",
	)

	checkin_map = {}
	for c in checkins:
		key = (c.employee, getdate(c.time))
		checkin_map.setdefault(key, []).append(c)
	return checkin_map


def get_shift_map():
	shifts = frappe.get_all("Shift Type", fields=["name", "start_time", "end_time"])
	return {s.name: s for s in shifts}


def get_source_label(checkin):
	"""
	Map the Employee Checkin source field (custom_source) to a display label:
	  - "Biometric"                -> "Biometric"
	  - empty / null               -> "Web"
	  - value containing "mobile"  -> "Mobile App"
	  - any other value            -> shown as-is
	"""
	if not checkin:
		return ""

	source = (checkin.get(CHECKIN_SOURCE_FIELD) or "").strip()
	if not source:
		return "Web"
	if source.lower() == "biometric":
		return "Biometric"
	if "mobile" in source.lower():
		return "Mobile App"
	return source


def get_first_in_and_last_out(checkins):
	"""
	First IN log of the day  -> check-in
	Last OUT log of the day  -> check-out
	(checkins list is already sorted by time asc)
	"""
	in_logs = [c for c in checkins if c.log_type == "IN"]
	out_logs = [c for c in checkins if c.log_type == "OUT"]

	first_in = in_logs[0] if in_logs else None
	last_out = out_logs[-1] if out_logs else None
	return first_in, last_out


def yes_no(value):
	return "Yes" if value else "No"


def _flag_value_or_blank(flag_value, custom_value):
	"""
	Show the HH:MM late / early-out duration when either:
	  - the standard Attendance boolean flag (late_entry / early_exit) is set, OR
	  - the custom HH:MM value has been populated by the sync job.
	Otherwise return an empty string. Prefer the HH:MM string over the flag
	so the report always displays how late / how early, not a 'Yes' label.
	"""
	if custom_value:
		return str(custom_value)
	if flag_value:
		return "Yes"
	return ""


def get_attendance_status_label(status):
	"""Map Attendance status to the single 'Attendance' column value."""
	if not status:
		return ""
	if status == "On Leave":
		return "Leave"
	return status  # Present / Absent / Half Day / Work From Home


def fmt_time(value):
	"""Format datetime / timedelta / time to HH:MM string."""
	if value is None:
		return ""
	# timedelta (Shift Type start/end times come as timedelta)
	if hasattr(value, "total_seconds") and not hasattr(value, "hour"):
		total_minutes = int(value.total_seconds() // 60)
		return f"{total_minutes // 60:02d}:{total_minutes % 60:02d}"
	# datetime / time
	try:
		return value.strftime("%H:%M")
	except Exception:
		return str(value)


def fmt_working_hours(in_time, out_time):
	"""Return the HH:MM duration between check-in and check-out.

	Returns an empty string if either timestamp is missing so the on-screen
	render and the Excel export stay consistent with the rest of the row.
	"""
	if not in_time or not out_time:
		return ""
	try:
		delta = out_time - in_time
	except TypeError:
		return ""
	total_minutes = int(delta.total_seconds() // 60)
	if total_minutes < 0:
		# OUT before IN (e.g. night shift crossing midnight) — add a day
		total_minutes += 24 * 60
	return f"{total_minutes // 60:02d}:{total_minutes % 60:02d}"


def build_row(att, checkins, shift_map):
	# CHANGED: takes the attendance record directly (it already carries the
	# joined Employee / Branch / HOD values). Every row has an attendance record.

	# First IN = check-in, Last OUT = check-out (from Employee Checkin)
	first_in, last_out = get_first_in_and_last_out(checkins)

	check_in_source = get_source_label(first_in)
	check_out_source = get_source_label(last_out)

	# Check-in / check-out times: prefer the actual checkin logs,
	# fall back to Attendance in/out time if no logs exist
	in_time = first_in.time if first_in else att.in_time
	out_time = last_out.time if last_out else att.out_time

	# Shift: prefer the shift on the attendance record, fall back to employee default
	shift_name = att.shift or att.default_shift
	shift = shift_map.get(shift_name)
	shift_time = (
		f"{fmt_time(shift.start_time)} - {fmt_time(shift.end_time)}" if shift else ""
	)

	# Late punch / early out: when the HRMS flag is true, replace the "Yes"
	# label with the HH:MM value stored in the custom field. When the flag is
	# false (or the custom value is empty), the column stays blank.
	late_punch = _flag_value_or_blank(att.late_entry, att.custom_late_arrival_minutes)
	early_out = _flag_value_or_blank(att.early_exit, att.custom_early_out_minutes)

	# Missed punch: true when either no IN or no OUT was logged for the day,
	# or the day is marked as On Leave (no expectation of either punch).
	has_in = bool(first_in) or bool(att.in_time)
	has_out = bool(last_out) or bool(att.out_time)
	is_on_leave = att.status == "On Leave"
	missed_punch = is_on_leave or (not has_in) or (not has_out)

	return {
		"ward": att.ward,
		"branch": att.branch,
		"employee": att.employee,
		"employee_name": att.employee_name,
		"department": att.department,
		"attendance_date": att.attendance_date,
		"hod_name": att.hod_name,
		"check_in_source": check_in_source,
		"check_out_source": check_out_source,
		"attendance": get_attendance_status_label(att.status),
		"shift_time": shift_time,
		"check_in_time": fmt_time(in_time),
		"check_out_time": fmt_time(out_time),
		"working_hours": fmt_working_hours(in_time, out_time),
		"late_punch": late_punch,
		"early_out": early_out,
		"missed_punch": yes_no(missed_punch),
	}


# ---------------------------------------------------------------------------
# Formatted Excel download (matches the on-screen table)
# ---------------------------------------------------------------------------

@frappe.whitelist()
def get_ward_options(txt=""):
	"""Distinct ward values from Branch.custom_ward for the Ward MultiSelectList.

	Restricted to wards belonging to branches the current user is permitted to view.
	"""
	permitted_branches = _get_permitted_branches()
	if permitted_branches is None:
		permitted_branches = [b.name for b in frappe.get_all("Branch", order_by="name asc")]

	filters = {"name": ("in", permitted_branches)}
	if txt:
		filters[BRANCH_WARD_FIELD] = ["like", f"%{txt}%"]

	wards = frappe.get_all(
		"Branch",
		filters=filters,
		pluck=BRANCH_WARD_FIELD,
		distinct=True,
	)
	return sorted({w for w in wards if w})


@frappe.whitelist()
def get_branch_options(txt="", wards=None):
	"""Branch MultiSelectList options, narrowed by selected Ward(s) and by the
	user's permitted branches."""
	permitted_branches = _get_permitted_branches()
	if permitted_branches is None:
		permitted_branches = [b.name for b in frappe.get_all("Branch", order_by="name asc")]

	filters = {"name": ("in", permitted_branches)}
	if txt:
		filters["name"] = ["in", permitted_branches]  # keep permitted scope
		# additional text filter is applied post-fetch
	ward_list = _as_list(wards)
	if ward_list:
		filters[BRANCH_WARD_FIELD] = ["in", ward_list]

	branches = frappe.get_all(
		"Branch",
		filters=filters,
		fields=["name"],
		order_by="name asc",
	)

	# Apply typeahead text filter on top of the permitted / ward-narrowed set
	if txt:
		needle = txt.strip().lower()
		branches = [b for b in branches if needle in (b.name or "").lower()]

	return branches


@frappe.whitelist()
def get_department_query(wards=None, branches=None):
	"""Return department names reachable from the selected wards/branches AND
	the user's permitted employees.

	Used by the Attendance Source Report filter's get_query to populate the
	Department dropdown options dynamically.

	Args:
	    wards:    JSON-encoded list of ward names (custom_ward on Branch)
	    branches: JSON-encoded list of branch names

	Returns:
	    list[dict] with `name` for the Department Link field.
	"""
	ward_list = _as_list(wards)
	branch_list = _as_list(branches)

	# Resolve ward -> branch list when only wards are selected
	if ward_list and not branch_list:
		branch_list = frappe.get_all(
			"Branch",
			filters={BRANCH_WARD_FIELD: ["in", ward_list]},
			pluck="name",
		) or []

	# Intersect with permitted branches
	permitted_branches = _get_permitted_branches()
	if permitted_branches is None:
		permitted_branches = [b.name for b in frappe.get_all("Branch", order_by="name asc")]

	if branch_list:
		branch_list = [b for b in branch_list if b in permitted_branches]
	else:
		branch_list = list(permitted_branches)

	if not branch_list:
		return []

	# Departments are derived from permitted Employees (not raw Employee table)
	# so users without standard read access to all Employees see only their own.
	permitted_employees = frappe.get_list("Employee", pluck="name")
	if not permitted_employees:
		return [d.name for d in frappe.get_all("Department", order_by="name asc")]

	dept_names = frappe.get_all(
		"Employee",
		filters={
			"name": ("in", permitted_employees),
			"branch": ("in", branch_list),
		},
		pluck="department",
		distinct=True,
	)
	dept_names = [d for d in dept_names if d]
	return [{"name": d} for d in sorted(set(dept_names))]

# ---------------------------------------------------------------------------
# Permission-aware filter options (mirrors Effective Attendance Report)code
# ---------------------------------------------------------------------------

def _get_explicit_branch_permissions():
	return frappe.get_all(
		"User Permission",
		filters={"user": frappe.session.user, "allow": "Branch"},
		pluck="for_value",
	)


def _get_derived_branch_permissions():
	# Respects standard Frappe perms + any User Permission rows with allow="Employee".
	permitted_employees = frappe.get_list("Employee", pluck="name")
	if not permitted_employees:
		return []

	return frappe.get_all(
		"Employee",
		filters={"name": ("in", permitted_employees)},
		pluck="branch",
	)


def _get_permitted_branches():
	"""Return the union of branches the current user can see.

	Returns None when no permission rows exist (admin fallback — caller should
	use all branches). Returns a sorted list otherwise.
	"""
	explicit = _get_explicit_branch_permissions()
	derived = _get_derived_branch_permissions()

	combined = {b for b in (explicit + derived) if b}

	if not combined:
		return None

	return sorted(combined)




@frappe.whitelist()
def download_excel(filters=None):
	if isinstance(filters, str):
		filters = json.loads(filters)
	filters = frappe._dict(filters or {})
	validate_filters(filters)

	columns = get_columns()
	data = get_data(filters)

	from openpyxl import Workbook
	from openpyxl.styles import Alignment, Border, Font, Side
	from openpyxl.utils import get_column_letter

	wb = Workbook()
	ws = wb.active
	ws.title = "Attendance Source"

	thin = Side(style="thin")
	border = Border(left=thin, right=thin, top=thin, bottom=thin)
	header_font = Font(bold=True, color="FFFFFF")
	header_fill = None
	try:
		from openpyxl.styles import PatternFill
		header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
	except Exception:
		header_fill = None
	center = Alignment(horizontal="center", vertical="center", wrap_text=True)
	left_align = Alignment(horizontal="left", vertical="center")

	# Header row (frozen)
	for col_idx, col in enumerate(columns, start=1):
		cell = ws.cell(row=1, column=col_idx, value=col["label"])
		cell.font = header_font
		cell.alignment = center
		cell.border = border
		if header_fill is not None:
			cell.fill = header_fill
	ws.row_dimensions[1].height = 28
	ws.freeze_panes = "A2"

	# Date-format map: only the Attendance Date column gets a real date format.
	date_fieldnames = {"attendance_date"}

	# Data rows
	for row_idx, row in enumerate(data, start=2):
		for col_idx, col in enumerate(columns, start=1):
			value = row.get(col["fieldname"], "")
			if col["fieldname"] in date_fieldnames and value:
				value = getdate(value)
			cell = ws.cell(row=row_idx, column=col_idx, value=value)
			cell.border = border
			# Centre-align short fields; left-align long text fields
			if col["fieldname"] in {"employee_name", "hod_name", "ward"}:
				cell.alignment = left_align
			else:
				cell.alignment = center
			if col["fieldname"] in date_fieldnames:
				cell.number_format = "dd-MM-yyyy"
			elif col["fieldname"] in {"check_in_time", "check_out_time"}:
				cell.number_format = "HH:MM"

	# Column widths from report definition (clamp to a reasonable max)
	for col_idx, col in enumerate(columns, start=1):
		width = col.get("width") or 100
		ws.column_dimensions[get_column_letter(col_idx)].width = min(max(int(width / 7), 10), 40)

	# Landscape print + fit-to-width
	ws.page_setup.orientation = "landscape"
	ws.page_setup.fitToWidth = 1
	ws.page_setup.fitToHeight = 0
	ws.print_options.horizontalCentered = True

	buffer = BytesIO()
	wb.save(buffer)

	frappe.local.response.filename = (
		f"Attendance Source - {formatdate(filters.from_date, 'dd-MM-yyyy')}"
		f" to {formatdate(filters.to_date, 'dd-MM-yyyy')}.xlsx"
	)
	frappe.local.response.filecontent = buffer.getvalue()
	frappe.local.response.type = "binary"