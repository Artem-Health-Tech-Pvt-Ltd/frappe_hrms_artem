# Copyright (c) 2026
# Effective of Contractual Staff - server side logic

BRANCH_WARD_FIELD = "custom_ward"
EMPLOYMENT_TYPE = "Contract"

import json
from io import BytesIO

import frappe
from frappe import _
from frappe.utils import add_days, date_diff, formatdate, getdate

from hrms.hr.doctype.leave_application.leave_application import get_leave_details


def execute(filters=None):
    filters = frappe._dict(filters or {})
    filters.branch = _as_list(filters.get("branch"))
    validate_filters(filters)

    period = get_period(filters)
    columns = get_columns(period)
    data = get_screen_data(filters, period)
    return columns, data


def _as_list(value):
    if not value:
        return []
    if isinstance(value, (list, tuple)):
        return [str(v) for v in value if v not in (None, "")]
    if isinstance(value, str):
        v = value.strip()
        if v.startswith("[") and v.endswith("]"):
            try:
                parsed = json.loads(v)
                if isinstance(parsed, list):
                    return [str(x) for x in parsed if x not in (None, "")]
            except Exception:
                pass
        return [v]
    return [str(value)]


# ---------------------------------------------------------------------------
# Strict Permitted Employee Extraction
# ---------------------------------------------------------------------------

def _get_permitted_employees():
    """
    Returns ONLY the exact employee records visible in the logged-in user's
    standard Employee List View.
    """
    if frappe.session.user == "Administrator":
        return None

    # Evaluates native Role & User Permissions strictly without manual expansion
    return frappe.get_all("Employee", pluck="name", ignore_permissions=False)


def _get_derived_branch_permissions():
    permitted_employees = _get_permitted_employees()
    if permitted_employees is None:
        return [b.name for b in frappe.get_all("Branch", order_by="name asc")]
    if not permitted_employees:
        return []

    branches = frappe.get_all(
        "Employee",
        filters={"name": ("in", permitted_employees)},
        pluck="branch",
        distinct=True,
    )
    return sorted({b for b in branches if b})


@frappe.whitelist()
def get_permitted_branches():
    return _get_derived_branch_permissions()


@frappe.whitelist()
def get_permitted_branches_for_multiselect(txt="", branches=None):
    permitted = get_permitted_branches()
    result = list(permitted)

    selected = _as_list(branches)
    if selected:
        result = [b for b in result if b in selected]

    if txt:
        needle = txt.strip().lower()
        result = [b for b in result if needle in b.lower()]

    return result


def validate_filters(filters):
    if not filters.get("from_date") or not filters.get("to_date"):
        frappe.throw(_("From Date and To Date are mandatory"))

    if getdate(filters.from_date) > getdate(filters.to_date):
        frappe.throw(_("From Date cannot be after To Date"))

    branch_values = _as_list(filters.get("branch"))
    if not branch_values:
        frappe.throw(_("Please select an Organization (Branch) from the filter to generate this report"))

    if date_diff(filters.to_date, filters.from_date) + 1 > 92:
        frappe.throw(_("Date range cannot exceed 92 days for this report"))


def get_period(filters):
    from_date = getdate(filters.from_date)
    to_date = getdate(filters.to_date)

    dates = []
    current = from_date
    while current <= to_date:
        dates.append(current)
        current = add_days(current, 1)

    single_month = from_date.month == to_date.month and from_date.year == to_date.year

    return frappe._dict(
        from_date=from_date,
        to_date=to_date,
        dates=dates,
        single_month=single_month,
        label=f"{formatdate(from_date, 'dd-MM-yyyy')} to {formatdate(to_date, 'dd-MM-yyyy')}",
    )


def get_day_label(date, period):
    if period.single_month:
        return str(date.day)
    return f"{date.day:02d}/{date.month:02d}"


def get_columns(period):
    columns = [
        {"label": _("Sr. No."), "fieldname": "sr_no", "fieldtype": "Data", "width": 60},
        {"label": _("Employee ID"), "fieldname": "employee", "fieldtype": "Data", "width": 110},
        {"label": _("Name"), "fieldname": "employee_name", "fieldtype": "Data", "width": 170},
        {"label": _("Designation"), "fieldname": "designation", "fieldtype": "Data", "width": 120},
        {"label": _("Joining Date"), "fieldname": "joining_date", "fieldtype": "Data", "width": 100},
        {"label": "", "fieldname": "row_label", "fieldtype": "Data", "width": 110},
    ]

    for idx, date in enumerate(period.dates, start=1):
        columns.append(
            {"label": get_day_label(date, period), "fieldname": f"d{idx}", "fieldtype": "Data", "width": 46}
        )

    columns.append({"label": _("Total Days"), "fieldname": "total_days", "fieldtype": "Data", "width": 80})
    columns.append({"label": _("Total Present Days"), "fieldname": "total_present_days", "fieldtype": "Data", "width": 100})
    columns.append({"label": _("Total Leaves Consumed"), "fieldname": "total_leaves_consumed", "fieldtype": "Int", "width": 110})
    columns.append({"label": _("Total Leaves Available"), "fieldname": "total_leaves_available", "fieldtype": "Int", "width": 110})

    return columns


def status_letter(status):
    if status in ("Present", "Half Day", "Work From Home"):
        return "P"
    if status == "Absent":
        return "A"
    if status == "On Leave":
        return "L"
    return ""


def format_hhmm(hours):
    if hours is None or hours == "":
        return ""
    if hasattr(hours, "total_seconds") and not hasattr(hours, "hour"):
        total_minutes = int(round(hours.total_seconds() / 60))
        return f"{total_minutes // 60:02d}:{total_minutes % 60:02d}"
    if hasattr(hours, "hour"):
        return f"{hours.hour:02d}:{hours.minute:02d}"
    try:
        total_minutes = int(round(float(hours) * 60))
        return f"{total_minutes // 60:02d}:{total_minutes % 60:02d}"
    except (TypeError, ValueError):
        return str(hours)


def get_employee_rows(filters, period):
    Attendance = frappe.qb.DocType("Attendance")
    Employee = frappe.qb.DocType("Employee")

    # Strictly limit to permitted employees first
    permitted_employees = _get_permitted_employees()
    if permitted_employees is not None and not permitted_employees:
        return []

    branch_list = filters.branch or get_permitted_branches()

    query = (
        frappe.qb.from_(Attendance)
        .inner_join(Employee)
        .on(Attendance.employee == Employee.name)
        .select(
            Attendance.employee,
            Attendance.employee_name,
            Attendance.attendance_date,
            Attendance.status,
            Attendance.working_hours,
            Employee.designation,
            Employee.date_of_joining,
        )
        .where(Attendance.docstatus == 1)
        .where(Attendance.attendance_date >= period.from_date)
        .where(Attendance.attendance_date <= period.to_date)
        .where(Employee.employment_type == EMPLOYMENT_TYPE)
    )

    if permitted_employees is not None:
        query = query.where(Employee.name.isin(permitted_employees))

    if branch_list:
        query = query.where(Employee.branch.isin(branch_list))

    query = query.orderby(Attendance.employee).orderby(Attendance.attendance_date)

    records = query.run(as_dict=True)

    emp_map = {}
    for rec in records:
        emp = emp_map.setdefault(
            rec.employee,
            {
                "employee": rec.employee,
                "employee_name": rec.employee_name,
                "designation": rec.designation,
                "date_of_joining": rec.date_of_joining,
                "days": {},
            },
        )
        hours = rec.working_hours or 0
        emp["days"][getdate(rec.attendance_date)] = {
            "letter": status_letter(rec.status),
            "hours": format_hhmm(hours) if hours else "",
        }

    employees = sorted(emp_map.values(), key=lambda e: (e["employee_name"] or ""))

    employee_ids = [e["employee"] for e in employees]
    balance_map = get_leave_balance_map(employee_ids, period.to_date) if employee_ids else {}

    for emp in employees:
        marked = [d for d in emp["days"].values() if d["letter"]]
        emp["total_days"] = len(marked)
        emp["total_present_days"] = len([d for d in marked if d["letter"] == "P"])
        emp["total_leaves_consumed"] = len([d for d in marked if d["letter"] == "L"])
        emp["total_leaves_available"] = balance_map.get(emp["employee"], 0)

    return employees


def get_leave_balance_map(employee_ids, date):
    balance_map = {}
    frappe.flags.ignore_permissions = True

    try:
        for employee in employee_ids:
            try:
                leave_details = get_leave_details(employee, date)
                balance_map[employee] = sum(
                    allocation.get("remaining_leaves") or 0
                    for allocation in leave_details.get("leave_allocation", {}).values()
                )
            except Exception:
                balance_map[employee] = 0
    finally:
        frappe.clear_messages()
        frappe.flags.ignore_permissions = False

    return balance_map


def get_screen_data(filters, period):
    employees = get_employee_rows(filters, period)

    data = []
    for sr_no, emp in enumerate(employees, start=1):
        pa_row = {
            "sr_no": str(sr_no),
            "employee": emp["employee"],
            "employee_name": emp["employee_name"],
            "designation": emp["designation"],
            "joining_date": formatdate(emp["date_of_joining"], "dd-MM-yyyy") if emp["date_of_joining"] else "",
            "row_label": "P/A",
            "total_days": str(emp["total_days"]),
            "total_present_days": str(emp["total_present_days"]),
            "total_leaves_consumed": emp["total_leaves_consumed"],
            "total_leaves_available": emp["total_leaves_available"],
        }
        hours_row = {
            "sr_no": "",
            "employee": "",
            "employee_name": "",
            "designation": "",
            "joining_date": "",
            "row_label": "Working Hours",
            "total_days": "",
            "total_present_days": "",
            "total_leaves_consumed": "",
            "total_leaves_available": "",
        }

        for idx, date in enumerate(period.dates, start=1):
            day_data = emp["days"].get(date)
            pa_row[f"d{idx}"] = day_data["letter"] if day_data else ""
            hours_row[f"d{idx}"] = day_data["hours"] if day_data else ""

        data.append(pa_row)
        data.append(hours_row)

    return data


@frappe.whitelist()
def download_excel(filters=None):
    if isinstance(filters, str):
        filters = json.loads(filters)
    filters = frappe._dict(filters or {})
    filters.branch = _as_list(filters.get("branch"))
    validate_filters(filters)

    period = get_period(filters)
    wb = build_workbook(filters, period)

    buffer = BytesIO()
    wb.save(buffer)

    frappe.local.response.filename = f"Effective of Contractual Staff - {period.label}.xlsx"
    frappe.local.response.filecontent = buffer.getvalue()
    frappe.local.response.type = "binary"


def build_workbook(filters, period):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.properties import PageSetupProperties

    employees = get_employee_rows(filters, period)

    branch_names = filters.branch or get_permitted_branches()
    wards = []
    if branch_names:
        rows = frappe.get_all(
            "Branch",
            filters={"name": ("in", branch_names)},
            pluck=BRANCH_WARD_FIELD,
        )
        wards = sorted({w for w in rows if w})
    ward = ", ".join(wards)

    title_branches = ", ".join(branch_names) if branch_names else ""

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center")
    bold = Font(bold=True)

    num_days = len(period.dates)
    first_day_col = 7
    total_days_col = first_day_col + num_days
    total_present_col = total_days_col + 1
    total_leaves_consumed_col = total_present_col + 1
    total_leaves_available_col = total_leaves_consumed_col + 1
    total_cols = total_leaves_available_col

    row = 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols)
    title_cell = ws.cell(row, 1, f"Effective of Contractual Staff at {title_branches} Dispensary/HBT")
    title_cell.font = Font(bold=True, size=13)
    title_cell.alignment = center
    ws.row_dimensions[row].height = 24
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.cell(row, 1, "To,").alignment = left_align
    row += 2

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ward_cell = ws.cell(row, 1, f"Ward - {ward}")
    ward_cell.font = Font(bold=True, underline="single")
    ward_cell.alignment = left_align

    ws.merge_cells(start_row=row, start_column=total_cols - 6, end_row=row, end_column=total_cols)
    period_cell = ws.cell(row, total_cols - 6, f"Period- {period.label}")
    period_cell.font = bold
    period_cell.alignment = Alignment(horizontal="right", vertical="center")
    row += 1

    header_row = row
    headers = ["Sr. No.", "Employee ID", "Name", "Designation", "Joining Date", ""]
    headers += [get_day_label(d, period) for d in period.dates]
    headers += ["Total Days", "Total Present Days", "Total Leaves Consumed", "Total Leaves Available"]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(header_row, col_idx, header)
        cell.font = bold
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[header_row].height = 30
    row += 1

    for sr_no, emp in enumerate(employees, start=1):
        top, bottom = row, row + 1

        identity_values = [
            sr_no,
            emp["employee"],
            emp["employee_name"],
            emp["designation"] or "",
            formatdate(emp["date_of_joining"], "dd-MM-yyyy") if emp["date_of_joining"] else "",
        ]
        for col_idx, value in enumerate(identity_values, start=1):
            ws.merge_cells(start_row=top, start_column=col_idx, end_row=bottom, end_column=col_idx)
            cell = ws.cell(top, col_idx, value)
            cell.alignment = center

        ws.cell(top, 6, "P/A").font = bold
        ws.cell(bottom, 6, "Working Hours").font = bold

        for idx, date in enumerate(period.dates):
            col = first_day_col + idx
            day_data = emp["days"].get(date)
            ws.cell(top, col, day_data["letter"] if day_data else "")
            ws.cell(bottom, col, day_data["hours"] if day_data else "")

        for col_idx, value in (
            (total_days_col, emp["total_days"]),
            (total_present_col, emp["total_present_days"]),
            (total_leaves_consumed_col, emp["total_leaves_consumed"]),
            (total_leaves_available_col, emp["total_leaves_available"]),
        ):
            ws.merge_cells(start_row=top, start_column=col_idx, end_row=bottom, end_column=col_idx)
            ws.cell(top, col_idx, value)

        for r in (top, bottom):
            for c in range(1, total_cols + 1):
                cell = ws.cell(r, c)
                cell.border = border
                if c >= first_day_col or c == 1:
                    cell.alignment = center
                elif c == 6:
                    cell.alignment = left_align

        row += 2

    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row + 1, end_column=5)
    sign_cell = ws.cell(row, 1, "Sign & stamp of MO/ Sr.MO")
    sign_cell.font = bold
    sign_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 20
    ws.row_dimensions[row + 1].height = 20

    name_width = max([len(e["employee_name"] or "") for e in employees] + [12]) + 2
    designation_width = max([len(e["designation"] or "") for e in employees] + [11]) + 2
    emp_id_width = max([len(e["employee"] or "") for e in employees] + [11]) + 2
    day_width = 4.5 if period.single_month else 6.5

    ws.column_dimensions["A"].width = 7
    ws.column_dimensions["B"].width = emp_id_width
    ws.column_dimensions["C"].width = name_width
    ws.column_dimensions["D"].width = designation_width
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 14
    for idx in range(num_days):
        ws.column_dimensions[get_column_letter(first_day_col + idx)].width = day_width
    ws.column_dimensions[get_column_letter(total_days_col)].width = 8
    ws.column_dimensions[get_column_letter(total_present_col)].width = 10
    ws.column_dimensions[get_column_letter(total_leaves_consumed_col)].width = 14
    ws.column_dimensions[get_column_letter(total_leaves_available_col)].width = 14

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.print_options.horizontalCentered = True
    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3
    ws.page_margins.top = 0.4
    ws.page_margins.bottom = 0.4

    return wb